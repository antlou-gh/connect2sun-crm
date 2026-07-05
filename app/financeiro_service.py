"""Serviços partilhados do módulo financeiro: importação e exportação Excel.

Princípios:
- **Ler/escrever SEMPRE por nome de cabeçalho**, nunca por posição de coluna
  (no ficheiro real o "NIF" está na coluna D, a seguir à "Descrição", o que
  desloca as colunas seguintes — o mapeamento posicional falharia).
- O export é um **snapshot unidirecional** (app → Excel). A app é a fonte de
  verdade; ficheiros exportados nunca são reimportados.
"""

import os
import re
import unicodedata
from datetime import date

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

from . import db
from .models import (
    Transacao, Client,
    ESTADOS, TIPOS_MOVIMENTO, CATEGORIAS, MESES,
)

SHEET_NAME = "BD Financeira 2026"
TABELA = "Tabela2"

# Caminho do ficheiro-modelo (cabeçalhos + estilos, sem dados) usado pelo export.
MODELO_PATH = os.path.join(
    os.path.dirname(__file__), "static", "templates_xlsx",
    "BD_Financeira_2026_modelo.xlsx",
)

# Cabeçalhos esperados na Tabela2 (textos exatos confirmados no ficheiro real).
CAB_NUMERO = "Nº de ordem"
CAB_DESCRICAO = "Descrição"
CAB_NIF = "NIF"
CAB_VALOR = "Valor total"
CAB_ENTIDADE = "Entidade emissora"
CAB_FACTURA = "Nº de factura"
CAB_SIVA = "S/ IVA"
CAB_IVA = "IVA"
CAB_IVA_PCT = "IVA %"
CAB_DIA = "Dia"
CAB_MES = "Mês"
CAB_TOTAL = "Total"
CAB_ESTADO = "Estado"
CAB_TIPO = "Tipo de movimento"

# Cabeçalhos obrigatórios para o import funcionar.
CABECALHOS_IMPORT = [
    CAB_NUMERO, CAB_DESCRICAO, CAB_NIF, CAB_VALOR, CAB_ENTIDADE, CAB_FACTURA,
    CAB_SIVA, CAB_IVA_PCT, CAB_DIA, CAB_MES, CAB_ESTADO, CAB_TIPO,
]


# ── Normalização ──────────────────────────────────────────────────────────────

def _normalizar(texto):
    """trim + minúsculas + sem acentos — para comparar cabeçalhos de forma robusta."""
    if texto is None:
        return ""
    s = str(texto).strip().lower()
    s = unicodedata.normalize("NFKD", s)
    return "".join(c for c in s if not unicodedata.combining(c))


def normalizar_nif(valor):
    """NIF → string canónica. No Excel vem como inteiro; no modelo é String(20).

    Converte com str(int(...)) quando numérico, tira espaços e um prefixo "PT".
    Devolve None se vazio.
    """
    if valor is None:
        return None
    if isinstance(valor, float):
        valor = int(valor)
    if isinstance(valor, int):
        s = str(valor)
    else:
        s = str(valor).strip().replace(" ", "")
        if s[:2].upper() == "PT":
            s = s[2:]
    s = s.strip()
    return s or None


def mes_para_numero(nome):
    """Nome do mês em português → 1..12. Tolera abreviaturas (ex.: 'Mar')."""
    if nome is None:
        return None
    alvo = _normalizar(nome)
    if not alvo:
        return None
    for i, m in enumerate(MESES, start=1):
        if _normalizar(m) == alvo:
            return i
    # Tolerância a abreviaturas: prefixo de 3 letras é único entre os 12 meses.
    if len(alvo) >= 3:
        for i, m in enumerate(MESES, start=1):
            if _normalizar(m).startswith(alvo[:3]):
                return i
    return None


# ── Localização da tabela / mapa de cabeçalhos ────────────────────────────────

def _selecionar_folha(wb):
    if SHEET_NAME in wb.sheetnames:
        return wb[SHEET_NAME]
    # Fallback: a folha que contém a Tabela2, senão a primeira.
    for ws in wb.worksheets:
        if TABELA in ws.tables:
            return ws
    return wb.worksheets[0]


def _limites_tabela(ws):
    """(min_col, header_row, max_col, max_row) da Tabela2; fallback B2:O? se ausente."""
    if TABELA in ws.tables:
        min_col, min_row, max_col, max_row = range_boundaries(ws.tables[TABELA].ref)
        return min_col, min_row, max_col, max_row
    # Sem tabela: assume cabeçalhos na linha 2, colunas B..O.
    return 2, 2, 15, ws.max_row


def _mapa_cabecalhos(ws, header_row, min_col, max_col):
    """{cabeçalho_normalizado: índice_coluna} lendo a linha de cabeçalhos."""
    mapa = {}
    for col in range(min_col, max_col + 1):
        valor = ws.cell(row=header_row, column=col).value
        chave = _normalizar(valor)
        if chave:
            mapa[chave] = col
    return mapa


def _exigir_cabecalhos(mapa, cabecalhos):
    em_falta = [c for c in cabecalhos if _normalizar(c) not in mapa]
    if em_falta:
        raise ValueError(
            "Cabeçalhos em falta na folha (lidos por nome, não por posição): "
            + ", ".join(em_falta)
        )


# ── Importação ────────────────────────────────────────────────────────────────

def importar_financeiro(caminho_xlsx, ano=2026):
    """Importa os movimentos do .xlsx para a tabela `transacoes`.

    - Lê colunas por NOME de cabeçalho.
    - `categoria` fica sempre NULL (categorização é feita depois na app).
    - `cliente_id` resolvido por NIF (NULL se vazio/sem correspondência).
    - Idempotente: `numero_ordem` já existente é saltado (não duplica).

    Devolve um dicionário-resumo.
    """
    wb = load_workbook(caminho_xlsx, data_only=True)
    ws = _selecionar_folha(wb)
    min_col, header_row, max_col, max_row = _limites_tabela(ws)
    mapa = _mapa_cabecalhos(ws, header_row, min_col, max_col)
    _exigir_cabecalhos(mapa, CABECALHOS_IMPORT)

    def ler(row, cabecalho):
        col = mapa[_normalizar(cabecalho)]
        return ws.cell(row=row, column=col).value

    # Índice de clientes por NIF normalizado (resolução rápida).
    clientes_por_nif = {}
    for c in Client.query.filter(Client.nif.isnot(None), Client.nif != "").all():
        nif = normalizar_nif(c.nif)
        if nif:
            clientes_por_nif.setdefault(nif, c.id)

    # Nºs de ordem já existentes — para idempotência.
    existentes = {
        n for (n,) in db.session.query(Transacao.numero_ordem)
        .filter(Transacao.numero_ordem.isnot(None)).all()
    }

    importados = ja_existentes = sem_cliente = ignoradas = 0
    avisos = []

    for row in range(header_row + 1, max_row + 1):
        descricao = ler(row, CAB_DESCRICAO)
        if descricao is None or str(descricao).strip() == "":
            continue  # linha vazia / fantasma (a tabela tem nºs de ordem pré-preenchidos)

        numero = ler(row, CAB_NUMERO)
        try:
            numero = int(numero)
        except (TypeError, ValueError):
            avisos.append(f"Linha {row}: Nº de ordem inválido ({numero!r}); ignorada.")
            ignoradas += 1
            continue

        if numero in existentes:
            ja_existentes += 1
            continue

        valor = ler(row, CAB_VALOR)
        try:
            valor = float(valor)
        except (TypeError, ValueError):
            avisos.append(f"Linha {row} (ordem {numero}): Valor total inválido; ignorada.")
            ignoradas += 1
            continue

        # Data a partir de Dia + Mês (texto pt) + ano.
        dia = ler(row, CAB_DIA)
        mes_num = mes_para_numero(ler(row, CAB_MES))
        try:
            data = date(ano, mes_num, int(dia))
        except (TypeError, ValueError):
            avisos.append(
                f"Linha {row} (ordem {numero}): Dia/Mês inválido "
                f"({dia!r}/{ler(row, CAB_MES)!r}); ignorada."
            )
            ignoradas += 1
            continue

        valor_siva = ler(row, CAB_SIVA)
        try:
            valor_siva = float(valor_siva)
        except (TypeError, ValueError):
            valor_siva = None
        # IVA: na folha é fórmula =ABS([Valor total])-[S/ IVA]; não guardar a string.
        iva = round(abs(valor) - valor_siva, 2) if valor_siva is not None else None

        iva_pct = ler(row, CAB_IVA_PCT)
        try:
            iva_pct = float(iva_pct)
        except (TypeError, ValueError):
            iva_pct = None

        cliente_id = clientes_por_nif.get(normalizar_nif(ler(row, CAB_NIF)))
        if cliente_id is None:
            sem_cliente += 1

        def texto(cabecalho):
            v = ler(row, cabecalho)
            return str(v).strip() if v is not None and str(v).strip() != "" else None

        t = Transacao(
            numero_ordem=numero,
            descricao=str(descricao).strip(),
            valor=valor,
            entidade_emissora=texto(CAB_ENTIDADE),
            num_factura=texto(CAB_FACTURA),
            valor_siva=valor_siva,
            iva=iva,
            iva_pct=iva_pct,
            data=data,
            estado=texto(CAB_ESTADO),
            tipo_movimento=texto(CAB_TIPO),
            categoria=None,  # sempre NULL na importação
            cliente_id=cliente_id,
        )
        db.session.add(t)
        existentes.add(numero)
        importados += 1

    db.session.commit()

    return {
        "importados": importados,
        "ja_existentes": ja_existentes,
        "sem_cliente": sem_cliente,
        "sem_categoria": importados,  # categoria é sempre NULL na importação
        "ignoradas": ignoradas,
        "avisos": avisos,
    }


# ── Exportação (snapshot app → Excel) ─────────────────────────────────────────

def gerar_export_financeiro(ano=2026):
    """Gera um Workbook (snapshot) das transações do `ano` no formato da folha.

    Parte do ficheiro-modelo (preserva Tabela2, cabeçalhos, estilos e larguras).
    Read-only sobre a BD. Escreve cada campo na coluna do respetivo cabeçalho.
    """
    if not os.path.exists(MODELO_PATH):
        raise FileNotFoundError(
            f"Ficheiro-modelo do export não encontrado: {MODELO_PATH}"
        )

    wb = load_workbook(MODELO_PATH)  # ler fresco a cada chamada (não mutar partilhado)
    ws = _selecionar_folha(wb)
    min_col, header_row, max_col, max_row = _limites_tabela(ws)
    mapa = _mapa_cabecalhos(ws, header_row, min_col, max_col)
    # Para o export precisamos também do cabeçalho "Total".
    _exigir_cabecalhos(mapa, CABECALHOS_IMPORT + [CAB_TOTAL])

    # O ficheiro-modelo traz um AutoFilter gravado (ex.: Estado = "Fechado") e
    # centenas de `row_dimensions[n].hidden = True` herdados de quando esse
    # filtro foi aplicado no Excel. Isso esconde linhas cujo Estado mudou
    # entretanto, e lotes inteiros de linhas novas. O export não deve herdar
    # nenhum estado de visibilidade — remove-se o filtro e força-se tudo
    # visível.
    if TABELA in ws.tables:
        ws.tables[TABELA].autoFilter = None
    for linha_dim in ws.row_dimensions.values():
        linha_dim.hidden = False

    transacoes = (
        Transacao.query
        .filter(db.extract("year", Transacao.data) == ano)
        .order_by(Transacao.numero_ordem.asc())
        .all()
    )

    FORMATO_MOEDA = '#,##0.00" €"'
    FORMATO_PCT = '0.00%'

    def escrever(linha, cabecalho, valor, number_format=None):
        col = mapa[_normalizar(cabecalho)]
        cell = ws.cell(row=linha, column=col, value=valor)
        if number_format:
            cell.number_format = number_format

    linha = header_row
    for t in transacoes:
        linha += 1
        ws.row_dimensions[linha].hidden = False
        nif = normalizar_nif(t.cliente.nif) if t.cliente else None
        mes_nome = MESES[t.data.month - 1] if t.data else None
        # round(2) imediatamente antes de escrever: evita ruído de vírgula
        # flutuante gravado na célula (ex.: 8945.7900000000009).
        valor = round(t.valor, 2) if t.valor is not None else None
        valor_siva = round(t.valor_siva, 2) if t.valor_siva is not None else None
        iva = round(t.iva, 2) if t.iva is not None else None
        total = round(abs(t.valor), 2) if t.valor is not None else None
        escrever(linha, CAB_NUMERO, t.numero_ordem)
        escrever(linha, CAB_DESCRICAO, t.descricao)
        escrever(linha, CAB_NIF, nif or "")
        escrever(linha, CAB_VALOR, valor, FORMATO_MOEDA)
        escrever(linha, CAB_ENTIDADE, t.entidade_emissora)
        escrever(linha, CAB_FACTURA, t.num_factura)
        escrever(linha, CAB_SIVA, valor_siva, FORMATO_MOEDA)
        escrever(linha, CAB_IVA, iva, FORMATO_MOEDA)  # valor literal
        escrever(linha, CAB_IVA_PCT, t.iva_pct, FORMATO_PCT)
        escrever(linha, CAB_DIA, t.data.day if t.data else None)
        escrever(linha, CAB_MES, mes_nome)           # texto português
        escrever(linha, CAB_TOTAL, total, FORMATO_MOEDA)  # literal, nunca fórmula
        escrever(linha, CAB_ESTADO, t.estado)
        escrever(linha, CAB_TIPO, t.tipo_movimento)
        # categoria NÃO é exportada (dimensão só da app; não existe na Tabela2).

    # Ajustar o intervalo da Tabela2 ao nº de linhas escritas.
    # Mínimo: cabeçalho + 1 linha (mantém a tabela válida mesmo sem dados).
    ultima = max(linha, header_row + 1)
    if TABELA in ws.tables:
        ref = (f"{get_column_letter(min_col)}{header_row}:"
               f"{get_column_letter(max_col)}{ultima}")
        ws.tables[TABELA].ref = ref

    return wb


# ── Criação de movimentos (regra de negócio partilhada app + /api/v1) ─────────
# Esta é a ÚNICA fonte de verdade da regra "criar uma transação válida". Tanto o
# endpoint humano (blueprints/financeiro.py) como a API de máquina
# (blueprints/api_v1.py) chamam criar_transacao_from_dict.

def _parse_date(value):
    if not value:
        return None
    try:
        return date.fromisoformat(value)
    except (ValueError, TypeError):
        return None


def _next_numero_ordem():
    atual = db.session.query(db.func.max(Transacao.numero_ordem)).scalar()
    return (atual or 0) + 1


def _validar_valor_fechado(valor, permitidos, label):
    """Devolve (valor_normalizado, erro). Vazio → None. Inválido → erro."""
    if valor in (None, ""):
        return None, None
    if valor not in permitidos:
        return None, f"{label} inválido: {valor!r}"
    return valor, None


def _aplicar_campos(t, body):
    """Aplica campos do body a uma Transacao; devolve lista de erros."""
    erros = []

    if "estado" in body:
        v, e = _validar_valor_fechado(body["estado"], ESTADOS, "Estado")
        if e:
            erros.append(e)
        else:
            t.estado = v
    if "tipo_movimento" in body:
        v, e = _validar_valor_fechado(body["tipo_movimento"], TIPOS_MOVIMENTO, "Tipo de movimento")
        if e:
            erros.append(e)
        else:
            t.tipo_movimento = v
    if "categoria" in body:
        v, e = _validar_valor_fechado(body["categoria"], CATEGORIAS, "Categoria")
        if e:
            erros.append(e)
        else:
            t.categoria = v

    if "data" in body:
        d = _parse_date(body["data"])
        if d is None:
            erros.append("Data inválida (esperado AAAA-MM-DD).")
        else:
            t.data = d

    if "cliente_id" in body:
        cid = body["cliente_id"]
        if cid in (None, ""):
            t.cliente_id = None
        else:
            try:
                cid = int(cid)
            except (TypeError, ValueError):
                erros.append("cliente_id inválido.")
                cid = None
            if cid is not None and not db.session.get(Client, cid):
                erros.append(f"Cliente {cid} não existe.")
            elif cid is not None:
                t.cliente_id = cid

    for campo in ("descricao", "entidade_emissora", "num_factura"):
        if campo in body:
            val = body[campo]
            setattr(t, campo, val.strip() if isinstance(val, str) and val.strip() else val or None)
    for campo in ("valor", "valor_siva", "iva", "iva_pct"):
        if campo in body:
            val = body[campo]
            if val in (None, ""):
                setattr(t, campo, None)
            else:
                try:
                    setattr(t, campo, float(val))
                except (TypeError, ValueError):
                    erros.append(f"{campo} inválido.")

    return erros


def criar_transacao_from_dict(body):
    """Cria (sem commit) uma Transacao a partir de um dict validado.

    Devolve (transacao, erros). Se erros não vazio, transacao é None.
    Valida obrigatórios (descricao, valor, data), aplica/valida campos contra
    as constantes fechadas, e atribui numero_ordem seguinte.
    NÃO faz db.session.commit() — quem chama decide quando fazer commit.

    Os erros de obrigatórios são devolvidos um a um (mesma semântica do endpoint
    humano original, que devolvia a primeira falha); os erros de _aplicar_campos
    vêm todos na lista. Quem chama junta-os com "; " na resposta JSON.

    `entidade_emissora` é obrigatória (à semelhança de `categoria`): sem ela,
    o movimento não é rastreável até ao fornecedor/cliente que o emitiu, e essa
    informação nunca mais é recuperável depois de perdida. `num_factura` fica
    opcional (nem todo o movimento tem documento associado, ex.: transferência
    bancária), mas incentivada.
    """
    if not (body.get("descricao") or "").strip():
        return None, ["Descrição é obrigatória."]
    if body.get("valor") in (None, ""):
        return None, ["Valor é obrigatório."]
    if not body.get("data"):
        return None, ["Data é obrigatória."]
    if not (body.get("entidade_emissora") or "").strip():
        return None, ["Entidade emissora é obrigatória."]

    t = Transacao(descricao="", valor=0.0, data=date.today())
    erros = _aplicar_campos(t, body)
    if erros:
        return None, erros
    t.numero_ordem = _next_numero_ordem()
    return t, []


# ── Relatório de lacunas: numero_factura / entidade_emissora em falta ────────
# Muitos movimentos antigos têm a informação no texto livre da Descrição
# (convenção "Pag./Rec. Inst <CÓDIGO> <Cliente> <detalhe>") mas não nos campos
# próprios. Isto gera um relatório de sugestões extraídas/correspondidas da
# Descrição — NUNCA escreve na BD por si só (ver `aplicar_lacunas_financeiro`).

PADRAO_FACTURA = re.compile(r"\b(?:FT|FR|FS|FA|NC)\s?[\dA-Z/.-]*\d\b", re.IGNORECASE)


def _vazio(valor):
    return valor is None or str(valor).strip() == ""


def _extrair_numero_factura(descricao):
    """Procura um padrão tipo FT1234 na Descrição. None se não encontrar."""
    m = PADRAO_FACTURA.search(descricao or "")
    return m.group(0).strip() if m else None


def _entidades_correspondentes(descricao, entidades_conhecidas):
    """Nomes conhecidos (conjunto fechado) que aparecem no texto da Descrição.

    `entidades_conhecidas` deve vir ordenada dos nomes mais longos para os mais
    curtos, para não perder um nome composto (ex.: "Aurensol, Lda") a favor de
    um prefixo comum a outro.
    """
    texto = descricao or ""
    return [e for e in entidades_conhecidas if e and e in texto]


def relatorio_lacunas_financeiro():
    """Identifica Transacao com numero_factura/entidade_emissora em falta e
    sugere valores extraídos/correspondidos da Descrição. Read-only.

    Devolve uma lista de dicts (um por movimento em falta) com:
    numero_ordem, descricao, num_factura_atual, num_factura_sugerido,
    entidade_atual, entidade_sugerida, entidade_candidatos (só quando
    ambíguo), estado ('ok' | 'sem correspondência' | 'ambíguo').
    """
    entidades_conhecidas = sorted(
        {
            e for (e,) in db.session.query(Transacao.entidade_emissora)
            .filter(Transacao.entidade_emissora.isnot(None), Transacao.entidade_emissora != "")
            .distinct().all()
        },
        key=len, reverse=True,
    )

    transacoes = (
        Transacao.query
        .filter(
            db.or_(
                Transacao.num_factura.is_(None), Transacao.num_factura == "",
                Transacao.entidade_emissora.is_(None), Transacao.entidade_emissora == "",
            )
        )
        .order_by(Transacao.numero_ordem.asc())
        .all()
    )

    linhas = []
    for t in transacoes:
        falta_factura = _vazio(t.num_factura)
        falta_entidade = _vazio(t.entidade_emissora)

        factura_sugerida = _extrair_numero_factura(t.descricao) if falta_factura else None

        candidatos_entidade = []
        entidade_sugerida = None
        if falta_entidade:
            candidatos_entidade = _entidades_correspondentes(t.descricao, entidades_conhecidas)
            if len(candidatos_entidade) == 1:
                entidade_sugerida = candidatos_entidade[0]

        if falta_entidade and len(candidatos_entidade) > 1:
            estado = "ambíguo"
        elif (falta_factura and not factura_sugerida) or (falta_entidade and not entidade_sugerida):
            estado = "sem correspondência"
        else:
            estado = "ok"

        linhas.append({
            "_id": t.id,
            "numero_ordem": t.numero_ordem,
            "descricao": t.descricao,
            "num_factura_atual": t.num_factura or "",
            "num_factura_sugerido": factura_sugerida or "",
            "entidade_atual": t.entidade_emissora or "",
            "entidade_sugerida": entidade_sugerida or "",
            "entidade_candidatos": ", ".join(candidatos_entidade) if len(candidatos_entidade) > 1 else "",
            "estado": estado,
        })
    return linhas


def aplicar_lacunas_financeiro(linhas):
    """Escreve na BD só as linhas com estado == 'ok' (sugestão única e sem
    ambiguidade). Linhas 'sem correspondência' e 'ambíguo' nunca são tocadas.

    Recebe a lista devolvida por `relatorio_lacunas_financeiro`. Faz commit.
    Devolve o nº de movimentos atualizados.
    """
    aplicados = 0
    for linha in linhas:
        if linha["estado"] != "ok":
            continue
        t = db.session.get(Transacao, linha["_id"])
        if t is None:
            continue
        alterado = False
        if linha["num_factura_sugerido"] and _vazio(t.num_factura):
            t.num_factura = linha["num_factura_sugerido"]
            alterado = True
        if linha["entidade_sugerida"] and _vazio(t.entidade_emissora):
            t.entidade_emissora = linha["entidade_sugerida"]
            alterado = True
        if alterado:
            aplicados += 1
    db.session.commit()
    return aplicados
